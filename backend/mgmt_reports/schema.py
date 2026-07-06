from datetime import timedelta
import base64
import io
import re

import graphene
from django.core.cache import cache
from django.utils import timezone
from graphene_django import DjangoObjectType
from graphql import GraphQLError

from identity.decorators import Roles, role_required
from ai_assistant.models import OrgAISettings, UserAISettings
from ai_assistant.ai_prompts import (
    build_prompt_context,
    build_prompt_result_pdf,
    execute_prompt_template,
)
from .cache_utils import mgmt_cave_stats_cache_key, mgmt_cave_stats_cache_timeout_seconds

from .models import AIPrompt, MonthlyReportSnapshot, ReportMailingList


class AIPromptType(DjangoObjectType):
    class Meta:
        model = AIPrompt
        fields = ('id', 'title', 'description', 'category', 'required_role', 'is_active', 'order')


# ---------------------------------------------------------------------------
# Lightweight stats types for the Reporting tab (Phase 2)
# ---------------------------------------------------------------------------

class StatusCount(graphene.ObjectType):
    status = graphene.String()
    count = graphene.Int()


class ACHStatsType(graphene.ObjectType):
    total = graphene.Int()
    created_last_30d = graphene.Int()
    by_status = graphene.List(StatusCount)


class AdvOpsStatsType(graphene.ObjectType):
    total = graphene.Int()
    created_last_30d = graphene.Int()
    by_status = graphene.List(StatusCount)
    by_priority = graphene.List(StatusCount)


class WorkbenchStatsType(graphene.ObjectType):
    total = graphene.Int()
    created_last_30d = graphene.Int()
    # DEPLOYED status means the rule has been pushed to GitHub and the target
    # platform — those workbenches are considered active.
    active_count = graphene.Int(description='Workbenches with status DEPLOYED (rule is active on target platform)')
    by_status = graphene.List(StatusCount)
    by_robustness = graphene.List(StatusCount)


class RulesKpiType(graphene.ObjectType):
    total = graphene.Int()
    created_last_30d = graphene.Int()
    active_count = graphene.Int()
    deprecated_count = graphene.Int()
    with_playbooks_count = graphene.Int()
    standalone_count = graphene.Int()


class MgmtCaveStatsType(graphene.ObjectType):
    ach = graphene.Field(ACHStatsType)
    advops = graphene.Field(AdvOpsStatsType)
    workbench = graphene.Field(WorkbenchStatsType)
    rules = graphene.Field(RulesKpiType)


# ---------------------------------------------------------------------------
# Phase 5: Historical trends
# ---------------------------------------------------------------------------

class MonthlySnapshotType(graphene.ObjectType):
    year = graphene.Int()
    month = graphene.Int()
    label = graphene.String(description='Human-readable label e.g. "Jun 2026"')
    stats = graphene.JSONString(description='Full stats payload for this month')


# ---------------------------------------------------------------------------
# Phase 5: Mailing list
# ---------------------------------------------------------------------------

class MailingListMemberType(graphene.ObjectType):
    id = graphene.ID()
    username = graphene.String()
    email = graphene.String()
    role = graphene.String()
    is_subscribed = graphene.Boolean()
    subscribed_at = graphene.DateTime()
    unsubscribed_at = graphene.DateTime()


def _payload_to_stats_type(payload):
    return MgmtCaveStatsType(
        ach=ACHStatsType(
            total=payload['ach']['total'],
            created_last_30d=payload['ach']['created_last_30d'],
            by_status=[StatusCount(status=i['status'], count=i['count']) for i in payload['ach']['by_status']],
        ),
        advops=AdvOpsStatsType(
            total=payload['advops']['total'],
            created_last_30d=payload['advops']['created_last_30d'],
            by_status=[StatusCount(status=i['status'], count=i['count']) for i in payload['advops']['by_status']],
            by_priority=[StatusCount(status=i['status'], count=i['count']) for i in payload['advops']['by_priority']],
        ),
        workbench=WorkbenchStatsType(
            total=payload['workbench']['total'],
            created_last_30d=payload['workbench']['created_last_30d'],
            active_count=payload['workbench']['active_count'],
            by_status=[StatusCount(status=i['status'], count=i['count']) for i in payload['workbench']['by_status']],
            by_robustness=[StatusCount(status=i['status'], count=i['count']) for i in payload['workbench']['by_robustness']],
        ),
        rules=RulesKpiType(
            total=payload['rules']['total'],
            created_last_30d=payload['rules']['created_last_30d'],
            active_count=payload['rules']['active_count'],
            deprecated_count=payload['rules']['deprecated_count'],
            with_playbooks_count=payload['rules']['with_playbooks_count'],
            standalone_count=payload['rules']['standalone_count'],
        ),
    )


def _require_org_context(user):
    org = getattr(user, 'organization', None)
    if org is None:
        raise GraphQLError(
            'MGMT Cave is organization-scoped. Assign this account to an organization first.'
        )
    return org


def _compute_mgmt_cave_stats_payload(org, last_30d):
    # --- ACH ---
    from ach.models import ACHAnalysis
    ach_qs = ACHAnalysis.objects.filter(owner__organization=org)
    ach_by_status = []
    for status in ['RESEARCH', 'FINISHED', 'APPROVED']:
        count = ach_qs.filter(status=status).count()
        ach_by_status.append({'status': status, 'count': count})

    # --- AdvOps ---
    from advops.models import ADVOPSReport
    advops_qs = ADVOPSReport.objects.filter(organization=org)
    advops_by_status = []
    for status in ['IDEA', 'RESEARCH', 'DEVELOPMENT', 'APPROVED', 'TESTING', 'DEPLOYED', 'TUNING']:
        count = advops_qs.filter(status=status).count()
        advops_by_status.append({'status': status, 'count': count})
    advops_by_priority = []
    for priority in ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW']:
        count = advops_qs.filter(priority=priority).count()
        advops_by_priority.append({'status': priority, 'count': count})

    # --- Workbench ---
    # DEPLOYED status means the rule has been pushed to GitHub and the target
    # platform, so DEPLOYED workbenches represent active rules.
    from playbooks.models import PlaybookGraph
    wb_qs = PlaybookGraph.objects.filter(organization=org)
    wb_by_status = []
    for status in ['IDEA', 'RESEARCH', 'DEVELOPMENT', 'REVIEW', 'APPROVED', 'TESTING', 'DEPLOYED', 'TUNING']:
        count = wb_qs.filter(status=status).count()
        wb_by_status.append({'status': status, 'count': count})
    wb_by_robustness = []
    for level in range(0, 6):
        count = wb_qs.filter(robustness_level=level).count()
        wb_by_robustness.append({'status': str(level), 'count': count})

    # --- Rules ---
    from rules.models import DetectionRule
    rules_qs = DetectionRule.objects.filter(organization=org)

    return {
        'ach': {
            'total': ach_qs.count(),
            'created_last_30d': ach_qs.filter(created_at__gte=last_30d).count(),
            'by_status': ach_by_status,
        },
        'advops': {
            'total': advops_qs.count(),
            'created_last_30d': advops_qs.filter(created_at__gte=last_30d).count(),
            'by_status': advops_by_status,
            'by_priority': advops_by_priority,
        },
        'workbench': {
            'total': wb_qs.count(),
            'created_last_30d': wb_qs.filter(created_at__gte=last_30d).count(),
            # DEPLOYED = rule is active on target platform
            'active_count': wb_qs.filter(status='DEPLOYED').count(),
            'by_status': wb_by_status,
            'by_robustness': wb_by_robustness,
        },
        'rules': {
            'total': rules_qs.count(),
            'created_last_30d': rules_qs.filter(created_at__gte=last_30d).count(),
            'active_count': rules_qs.filter(playbook__status='DEPLOYED').count(),
            'deprecated_count': rules_qs.filter(status__iexact='deprecated').count(),
            'with_playbooks_count': rules_qs.filter(playbook__isnull=False).count(),
            'standalone_count': rules_qs.filter(playbook__isnull=True).count(),
        },
    }


def _build_excel_report(stats_payload):
    """Build an Excel workbook from a stats payload dict. Returns bytes."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='1677FF')
    center = Alignment(horizontal='center')

    def _add_sheet(name, headers, rows):
        ws = wb.create_sheet(title=name)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].auto_size = True
        return ws

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Overview
    ach = stats_payload.get('ach', {})
    advops = stats_payload.get('advops', {})
    wb_data = stats_payload.get('workbench', {})
    rules = stats_payload.get('rules', {})
    _add_sheet('Overview', ['Domain', 'Total', 'Created (30d)'], [
        ['ACH Analyses', ach.get('total', 0), ach.get('created_last_30d', 0)],
        ['AdvOps Hunts', advops.get('total', 0), advops.get('created_last_30d', 0)],
        ['Detection Workbenches', wb_data.get('total', 0), wb_data.get('created_last_30d', 0)],
        ['Detection Rules', rules.get('total', 0), rules.get('created_last_30d', 0)],
    ])

    # ACH
    _add_sheet('ACH by Status', ['Status', 'Count'],
               [[s['status'], s['count']] for s in ach.get('by_status', [])])

    # AdvOps
    _add_sheet('AdvOps by Status', ['Status', 'Count'],
               [[s['status'], s['count']] for s in advops.get('by_status', [])])
    _add_sheet('AdvOps by Priority', ['Priority', 'Count'],
               [[s['status'], s['count']] for s in advops.get('by_priority', [])])

    # Workbench
    _add_sheet('Workbench by Status', ['Status', 'Count', 'Note'],
               [[s['status'], s['count'],
                 'Rule is active on target platform' if s['status'] == 'DEPLOYED' else '']
                for s in wb_data.get('by_status', [])])
    _add_sheet('Workbench by Robustness', ['Level', 'Count'],
               [[s['status'], s['count']] for s in wb_data.get('by_robustness', [])])

    # Rules
    _add_sheet('Rules KPIs', ['Metric', 'Count'], [
        ['Total Rules', rules.get('total', 0)],
        ['Active Rules', rules.get('active_count', 0)],
        ['Deprecated Rules', rules.get('deprecated_count', 0)],
        ['With Workbench', rules.get('with_playbooks_count', 0)],
        ['Standalone', rules.get('standalone_count', 0)],
        ['Created (30d)', rules.get('created_last_30d', 0)],
    ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class Query(graphene.ObjectType):
    ai_prompts = graphene.List(
        AIPromptType,
        description='Retrieve all active AI prompts for the management dashboard',
    )

    mgmt_cave_stats = graphene.Field(
        MgmtCaveStatsType,
        description='Aggregate statistics for the MGMT Cave reporting tab',
    )

    monthly_trends = graphene.List(
        MonthlySnapshotType,
        months=graphene.Int(default_value=6, description='Number of past months to return (max 24)'),
        description='Month-over-month statistics snapshots for trend analysis',
    )

    mailing_list_members = graphene.List(
        MailingListMemberType,
        description='Users on the monthly report mailing list (admin only)',
    )

    @role_required([Roles.ADMIN, Roles.REVIEWER])
    def resolve_ai_prompts(self, info):
        user = info.context.user
        _require_org_context(user)
        queryset = AIPrompt.objects.filter(is_active=True)

        if user.role == Roles.REVIEWER:
            queryset = queryset.filter(required_role='REVIEWER')

        return queryset

    @role_required([Roles.ADMIN, Roles.REVIEWER])
    def resolve_mgmt_cave_stats(self, info):
        user = info.context.user
        org = _require_org_context(user)
        now = timezone.now()
        last_30d = now - timedelta(days=30)
        cache_key = mgmt_cave_stats_cache_key(getattr(org, 'id', 'none'))
        cached_payload = cache.get(cache_key)
        if cached_payload is not None:
            return _payload_to_stats_type(cached_payload)

        payload = _compute_mgmt_cave_stats_payload(org, last_30d)
        cache.set(cache_key, payload, timeout=mgmt_cave_stats_cache_timeout_seconds())
        return _payload_to_stats_type(payload)

    @role_required([Roles.ADMIN, Roles.REVIEWER])
    def resolve_monthly_trends(self, info, months=6):
        import calendar
        user = info.context.user
        org = _require_org_context(user)
        months = max(1, min(months, 24))
        now = timezone.now()
        snapshots = MonthlyReportSnapshot.objects.filter(organization=org).order_by('-year', '-month')[:months]
        result = []
        for snap in reversed(list(snapshots)):
            month_name = calendar.month_abbr[snap.month]
            result.append(MonthlySnapshotType(
                year=snap.year,
                month=snap.month,
                label=f'{month_name} {snap.year}',
                stats=snap.stats_json,
            ))
        return result

    @role_required([Roles.ADMIN])
    def resolve_mailing_list_members(self, info):
        user = info.context.user
        org = _require_org_context(user)
        entries = ReportMailingList.objects.filter(organization=org).select_related('user')
        result = []
        for entry in entries:
            result.append(MailingListMemberType(
                id=str(entry.id),
                username=entry.user.username,
                email=entry.user.email,
                role=entry.user.role,
                is_subscribed=entry.is_subscribed,
                subscribed_at=entry.subscribed_at,
                unsubscribed_at=entry.unsubscribed_at,
            ))
        return result


def _get_effective_ai_settings(user_settings):
    if getattr(user_settings, 'use_org_ai', False):
        org = getattr(user_settings.user, 'organization', None)
        if org:
            try:
                org_settings = OrgAISettings.objects.select_related('shared_profile').get(organization=org)
                effective = org_settings.get_effective_settings()
                if getattr(effective, 'has_any_provider', False):
                    return effective
            except OrgAISettings.DoesNotExist:
                pass
    return user_settings


class ExecuteAIPrompt(graphene.Mutation):
    class Arguments:
        prompt_id = graphene.UUID(required=True)
        custom_input = graphene.String(required=False)
        custom_context = graphene.JSONString(required=False)

    success = graphene.Boolean()
    message = graphene.String()
    result_markdown = graphene.String()
    rendered_prompt = graphene.String()
    provider_used = graphene.String()

    @staticmethod
    @role_required([Roles.ADMIN, Roles.REVIEWER])
    def mutate(root, info, prompt_id, custom_input=None, custom_context=None):
        user = info.context.user
        _require_org_context(user)
        try:
            prompt = AIPrompt.objects.get(id=prompt_id, is_active=True)
        except AIPrompt.DoesNotExist:
            raise GraphQLError("Prompt not found or inactive.")

        if user.role == Roles.REVIEWER and prompt.required_role == Roles.ADMIN:
            raise GraphQLError("You do not have permission to execute this prompt.")

        try:
            settings = UserAISettings.objects.get(user=user)
        except UserAISettings.DoesNotExist:
            return ExecuteAIPrompt(
                success=False,
                message="Please configure AI Settings in your profile first.",
                result_markdown=None,
                rendered_prompt=None,
                provider_used="NONE",
            )

        effective_settings = _get_effective_ai_settings(settings)
        context = build_prompt_context(
            user=user,
            custom_input=custom_input,
            custom_context=custom_context,
        )

        try:
            rendered_prompt, result_markdown, provider = execute_prompt_template(
                user_settings=effective_settings,
                prompt_template=prompt.prompt_template,
                context=context,
                prompt_title=prompt.title,
            )
            return ExecuteAIPrompt(
                success=True,
                message="Prompt executed successfully.",
                result_markdown=result_markdown,
                rendered_prompt=rendered_prompt,
                provider_used=provider,
            )
        except Exception as exc:
            return ExecuteAIPrompt(
                success=False,
                message=f"Prompt execution failed: {exc}",
                result_markdown=None,
                rendered_prompt=None,
                provider_used="ERROR",
            )


class ExportAIPromptResultPdf(graphene.Mutation):
    class Arguments:
        title = graphene.String(required=True)
        result_markdown = graphene.String(required=True)

    success = graphene.Boolean()
    message = graphene.String()
    file_data = graphene.String()
    filename = graphene.String()
    content_type = graphene.String()

    @staticmethod
    @role_required([Roles.ADMIN, Roles.REVIEWER])
    def mutate(root, info, title, result_markdown):
        _require_org_context(info.context.user)
        try:
            pdf_bytes = build_prompt_result_pdf(title=title, markdown_text=result_markdown)
        except Exception as exc:
            return ExportAIPromptResultPdf(
                success=False,
                message=f"Failed to generate PDF: {exc}",
                file_data=None,
                filename=None,
                content_type=None,
            )

        safe_title = re.sub(r"[^a-zA-Z0-9_-]", "_", title.strip() or "ai_prompt_result")
        return ExportAIPromptResultPdf(
            success=True,
            message="PDF generated successfully.",
            file_data=base64.b64encode(pdf_bytes).decode("utf-8"),
            filename=f"{safe_title}.pdf",
            content_type="application/pdf",
        )


class ExportReportExcel(graphene.Mutation):
    """Export the current MGMT Cave statistics report as an Excel workbook."""

    class Arguments:
        sections = graphene.List(
            graphene.String,
            required=False,
            description='Optional list of section keys to include (ach, advops, workbench, rules). '
                        'Omit to include all sections.',
        )

    success = graphene.Boolean()
    message = graphene.String()
    file_data = graphene.String()
    filename = graphene.String()
    content_type = graphene.String()

    @staticmethod
    @role_required([Roles.ADMIN, Roles.REVIEWER])
    def mutate(root, info, sections=None):
        user = info.context.user
        org = _require_org_context(user)
        now = timezone.now()
        last_30d = now - timedelta(days=30)

        cache_key = mgmt_cave_stats_cache_key(getattr(org, 'id', 'none'))
        payload = cache.get(cache_key) or _compute_mgmt_cave_stats_payload(org, last_30d)

        if sections:
            allowed = {'ach', 'advops', 'workbench', 'rules'}
            filtered = {k: v for k, v in payload.items() if k in sections and k in allowed}
            if not filtered:
                return ExportReportExcel(success=False, message='No valid sections specified.',
                                         file_data=None, filename=None, content_type=None)
            payload = filtered

        try:
            xlsx_bytes = _build_excel_report(payload)
        except Exception as exc:
            return ExportReportExcel(
                success=False, message=f"Excel generation failed: {exc}",
                file_data=None, filename=None, content_type=None,
            )

        date_str = now.strftime('%Y-%m-%d')
        return ExportReportExcel(
            success=True,
            message="Excel report generated successfully.",
            file_data=base64.b64encode(xlsx_bytes).decode("utf-8"),
            filename=f"hefaistos_report_{date_str}.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class UpdateMailingListMember(graphene.Mutation):
    """Add or remove a user from the monthly report mailing list (admin only)."""

    class Arguments:
        username = graphene.String(required=True)
        subscribe = graphene.Boolean(required=True, description='True to subscribe, False to remove')

    success = graphene.Boolean()
    message = graphene.String()

    @staticmethod
    @role_required([Roles.ADMIN])
    def mutate(root, info, username, subscribe):
        from django.contrib.auth import get_user_model
        User = get_user_model()
        admin_user = info.context.user
        org = _require_org_context(admin_user)

        try:
            target_user = User.objects.get(username=username, organization=org)
        except User.DoesNotExist:
            return UpdateMailingListMember(success=False, message='User not found in your organization.')

        entry, _ = ReportMailingList.objects.get_or_create(
            user=target_user,
            defaults={'organization': org, 'is_subscribed': subscribe},
        )
        if entry.is_subscribed == subscribe:
            state = 'subscribed' if subscribe else 'unsubscribed'
            return UpdateMailingListMember(success=True, message=f'User is already {state}.')

        entry.is_subscribed = subscribe
        if not subscribe:
            entry.unsubscribed_at = timezone.now()
            entry.removed_by = admin_user
        else:
            entry.unsubscribed_at = None
            entry.removed_by = None
        entry.save(update_fields=['is_subscribed', 'unsubscribed_at', 'removed_by'])
        state = 'subscribed' if subscribe else 'removed from the mailing list'
        return UpdateMailingListMember(success=True, message=f'User has been {state}.')


class Mutation(graphene.ObjectType):
    execute_ai_prompt = ExecuteAIPrompt.Field()
    export_ai_prompt_result_pdf = ExportAIPromptResultPdf.Field()
    export_report_excel = ExportReportExcel.Field()
    update_mailing_list_member = UpdateMailingListMember.Field()


schema = graphene.Schema(query=Query, mutation=Mutation)
